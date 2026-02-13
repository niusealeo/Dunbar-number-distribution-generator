#!/usr/bin/env python3
"""
Build an XLSX workbook for N = 9..564 with columns p, z, sigma, using:
- Dunbar:   System F (skew-normal fit) vs Composite Fuzzy (piecewise A–E) + differences
- Janson:   System F (skew-normal fit) vs Composite Fuzzy (piecewise A–E) + differences
- Dunbar-Janson: System F (skew-normal fit to averaged bounds) vs M6 (winner-stitched + fuzzy) + differences
- Definitions: equations, parameters, and interpretation

Notes / conventions:
- mu is fixed at 147.8 for all regimes (as in our work).
- "z" is the *standard normal* z = Φ^{-1}(p).
- For System F (skew-normal), "sigma" is reported as the fitted *scale* ω (constant per regime).
- For Composite Fuzzy / M6, "sigma" is the model’s effective σ(p) from the underlying A–E systems
  (with split σL/σU for B and E).
- Differences are reported as: Δp = p_model2 - p_model1, Δz, Δsigma, and ΔN(p) at the same p (optional).

Requires:
    pip install openpyxl mpmath
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from typing import Callable, Dict, Tuple, List

import mpmath as mp
from statistics import NormalDist
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Global constants
# -----------------------------

MU = 147.8
ND = NormalDist()

N_MIN = 9
N_MAX = 564

# 95% equal-tail quantiles:
P_LO = 0.025
P_HI = 0.975

# Bounds used for skew-normal System F fits
# Dunbar: (literature values)
DUNBAR_L = 100.2
DUNBAR_U = 231.1

# Janson: (as you provided)
JANSON_L = 22.73
JANSON_U = 446.2

# Dunbar-Janson "average" bounds (used to define a DJ System F fit)
DJ_L = (DUNBAR_L + JANSON_L) / 2.0
DJ_U = (DUNBAR_U + JANSON_U) / 2.0


# -----------------------------
# A–E σ parameters (pre-averaging)
# -----------------------------

DUNBAR_SIG = {
    "A": {"sigma": 33.393},
    "C": {"sigma": 28.629},
    "D": {"sigma": 31.011},
    "B": {"sigma_L": 24.286, "sigma_U": 42.501},
    "E": {"sigma_L": 26.458, "sigma_U": 35.565},
}

JANSON_SIG = {
    "A": {"sigma": 108.03004630194182},
    "C": {"sigma": 76.0177544515727},
    "D": {"sigma": 92.02390037675727},
    "B": {"sigma_L": 63.812397057566606, "sigma_U": 152.24769554631703},
    "E": {"sigma_L": 69.91507575456966, "sigma_U": 114.13272499894487},
}

# Dunbar-Janson averaged σ (componentwise mean) for use in M3 (pairwise) & other DJ models
AVG_SIG = {}
for k in ["A", "C", "D"]:
    AVG_SIG[k] = {"sigma": (DUNBAR_SIG[k]["sigma"] + JANSON_SIG[k]["sigma"]) / 2.0}
for k in ["B", "E"]:
    AVG_SIG[k] = {
        "sigma_L": (DUNBAR_SIG[k]["sigma_L"] + JANSON_SIG[k]["sigma_L"]) / 2.0,
        "sigma_U": (DUNBAR_SIG[k]["sigma_U"] + JANSON_SIG[k]["sigma_U"]) / 2.0,
    }


# -----------------------------
# Helpers: Normal, Owen's T, Skew-normal
# -----------------------------

def phi(x: mp.mpf) -> mp.mpf:
    return mp.exp(-x * x / 2.0) / mp.sqrt(2.0 * mp.pi)

def Phi(x: mp.mpf) -> mp.mpf:
    # mpmath erf-based normal CDF
    return 0.5 * (1.0 + mp.erf(x / mp.sqrt(2.0)))

def owens_t(h: mp.mpf, a: mp.mpf) -> mp.mpf:
    """
    Owen's T(h, a) = (1/(2π)) ∫_0^a exp(-0.5 h^2 (1+t^2)) / (1+t^2) dt
    """
    if a == 0:
        return mp.mpf("0")
    sgn = 1
    if a < 0:
        a = -a
        sgn = -1
    integrand = lambda t: mp.exp(-0.5 * h * h * (1 + t * t)) / (1 + t * t)
    val = (1.0 / (2.0 * mp.pi)) * mp.quad(integrand, [0, a])
    return sgn * val

def skewnorm_cdf_std(z: mp.mpf, alpha: mp.mpf) -> mp.mpf:
    """
    Standard skew-normal CDF for z with shape alpha:
        F(z) = Φ(z) - 2*T(z, alpha)
    """
    return Phi(z) - 2.0 * owens_t(z, alpha)

def skewnorm_mean_shift(alpha: float) -> float:
    """
    For standard SN(0,1,alpha), mean = delta * sqrt(2/pi),
    where delta = alpha / sqrt(1+alpha^2)
    """
    a = float(alpha)
    delta = a / math.sqrt(1.0 + a * a)
    return delta * math.sqrt(2.0 / math.pi)

@dataclass
class SkewNormalFit:
    """
    Skew-normal parameters with mean constrained to MU:
        X = xi + omega * Z,   Z ~ SN(0,1,alpha)
        E[X] = MU => xi = MU - omega * mean(Z)
    """
    omega: float
    alpha: float

    @property
    def xi(self) -> float:
        m = skewnorm_mean_shift(self.alpha)
        return MU - self.omega * m

    def cdf(self, x: float) -> float:
        z = (x - self.xi) / self.omega
        return float(skewnorm_cdf_std(mp.mpf(z), mp.mpf(self.alpha)))

    def inv_cdf(self, p: float) -> float:
        # bisection on z then transform to x
        lo, hi = -20.0, 20.0
        for _ in range(100):
            mid = (lo + hi) / 2.0
            val = float(skewnorm_cdf_std(mp.mpf(mid), mp.mpf(self.alpha)))
            if val < p:
                lo = mid
            else:
                hi = mid
        z = (lo + hi) / 2.0
        return self.xi + self.omega * z

def fit_skewnormal_from_bounds(L: float, U: float, p_lo=P_LO, p_hi=P_HI) -> SkewNormalFit:
    """
    Fit (omega, alpha) such that:
        Q(p_lo)=L, Q(p_hi)=U, mean fixed at MU
    """
    # initial guesses: near symmetric
    omega0 = (U - L) / (2.0 * 1.95996398454005)  # ~ 95% width / (2*z_0.975)
    alpha0 = 0.0

    def eqs(vars_):
        omega, alpha = vars_
        fit = SkewNormalFit(float(omega), float(alpha))
        return (fit.inv_cdf(p_lo) - L, fit.inv_cdf(p_hi) - U)

    # Use mp.findroot on omega, alpha
    # We clamp omega positive by solving for log_omega
    def f1(logw, a):
        w = mp.e**logw
        fit = SkewNormalFit(float(w), float(a))
        return mp.mpf(fit.inv_cdf(p_lo) - L)

    def f2(logw, a):
        w = mp.e**logw
        fit = SkewNormalFit(float(w), float(a))
        return mp.mpf(fit.inv_cdf(p_hi) - U)

    logw0 = math.log(max(omega0, 1e-6))
    a0 = alpha0
    # second starting point for robustness
    logw1 = logw0 + 0.05
    a1 = 2.0

    try:
        sol = mp.findroot((f1, f2), (logw0, a0), (logw1, a1), tol=1e-10, maxsteps=50)
        logw_sol, a_sol = float(sol[0]), float(sol[1])
    except Exception:
        # fallback: broader initial alpha
        sol = mp.findroot((f1, f2), (logw0, 5.0), (logw0 + 0.05, 10.0), tol=1e-10, maxsteps=80)
        logw_sol, a_sol = float(sol[0]), float(sol[1])

    omega_sol = math.exp(logw_sol)
    return SkewNormalFit(omega=omega_sol, alpha=a_sol)


# -----------------------------
# A–E system quantiles and σ(p)
# -----------------------------

def sigma_side(sig_params: Dict, sys: str, p: float) -> float:
    if sys in ("B", "E"):
        return sig_params[sys]["sigma_L"] if p < 0.5 else sig_params[sys]["sigma_U"]
    return sig_params[sys]["sigma"]

def z_std(p: float) -> float:
    return ND.inv_cdf(p)

def Q_system(sig_params: Dict, sys: str, p: float) -> float:
    return MU + sigma_side(sig_params, sys, p) * z_std(p)


# -----------------------------
# Composite fuzzy logic systems (your updated smoothing around 0.5)
# -----------------------------

def Dunbar_composite_Q(p: float) -> float:
    # Dunbar frame:
    # C ≤ 1/4 < 12(C(1/3-p)+D(p-1/4)) < 1/3 ≤ 6(D(1/2-p)+E(p-1/3)) < 1/2 < E
    if p <= 1/4:
        return Q_system(DUNBAR_SIG, "C", p)
    if p < 1/3:
        wC = 12.0 * (1/3 - p)
        wD = 12.0 * (p - 1/4)
        return wC * Q_system(DUNBAR_SIG, "C", p) + wD * Q_system(DUNBAR_SIG, "D", p)
    if p < 1/2:
        wD = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wD * Q_system(DUNBAR_SIG, "D", p) + wE * Q_system(DUNBAR_SIG, "E", p)
    return Q_system(DUNBAR_SIG, "E", p)

def Dunbar_composite_sigma(p: float) -> float:
    if p <= 1/4:
        return sigma_side(DUNBAR_SIG, "C", p)
    if p < 1/3:
        wC = 12.0 * (1/3 - p)
        wD = 12.0 * (p - 1/4)
        return wC * sigma_side(DUNBAR_SIG, "C", p) + wD * sigma_side(DUNBAR_SIG, "D", p)
    if p < 1/2:
        wD = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wD * sigma_side(DUNBAR_SIG, "D", p) + wE * sigma_side(DUNBAR_SIG, "E", p)
    return sigma_side(DUNBAR_SIG, "E", p)

def Janson_composite_Q(p: float) -> float:
    # Janson frame:
    # D ≤ 1/6 < 12(D(1/4-p)+A(p-1/6)) < 1/4 ≤ 6(A(1/2-p)+E(p-1/3)) < 1/2 < E ≤ 2/3 < 12(E(3/4-p)+B(p-2/3)) < 3/4 ≤ B
    if p <= 1/6:
        return Q_system(JANSON_SIG, "D", p)
    if p < 1/4:
        wD = 12.0 * (1/4 - p)
        wA = 12.0 * (p - 1/6)
        return wD * Q_system(JANSON_SIG, "D", p) + wA * Q_system(JANSON_SIG, "A", p)
    if p < 1/3:
        return Q_system(JANSON_SIG, "A", p)
    if p < 1/2:
        wA = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wA * Q_system(JANSON_SIG, "A", p) + wE * Q_system(JANSON_SIG, "E", p)
    if p <= 2/3:
        return Q_system(JANSON_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * Q_system(JANSON_SIG, "E", p) + wB * Q_system(JANSON_SIG, "B", p)
    return Q_system(JANSON_SIG, "B", p)

def Janson_composite_sigma(p: float) -> float:
    if p <= 1/6:
        return sigma_side(JANSON_SIG, "D", p)
    if p < 1/4:
        wD = 12.0 * (1/4 - p)
        wA = 12.0 * (p - 1/6)
        return wD * sigma_side(JANSON_SIG, "D", p) + wA * sigma_side(JANSON_SIG, "A", p)
    if p < 1/3:
        return sigma_side(JANSON_SIG, "A", p)
    if p < 1/2:
        wA = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wA * sigma_side(JANSON_SIG, "A", p) + wE * sigma_side(JANSON_SIG, "E", p)
    if p <= 2/3:
        return sigma_side(JANSON_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * sigma_side(JANSON_SIG, "E", p) + wB * sigma_side(JANSON_SIG, "B", p)
    return sigma_side(JANSON_SIG, "B", p)


# -----------------------------
# Dunbar-Janson mean regime models M1..M6 (with your latest renames)
# -----------------------------

# Pairwise-constructed model (this is NEW M3):
# knots/pairs used earlier:
PAIR_BD = ("B", "D")
PAIR_AE = ("A", "E")
PAIR_AD = ("A", "D")
PAIR_AB = ("A", "B")
PAIR_BE = ("B", "E")

def P_pair(pair: Tuple[str, str], p: float) -> float:
    a, b = pair
    return 0.5 * (Q_system(AVG_SIG, a, p) + Q_system(AVG_SIG, b, p))

def P_pair_sigma(pair: Tuple[str, str], p: float) -> float:
    a, b = pair
    return 0.5 * (sigma_side(AVG_SIG, a, p) + sigma_side(AVG_SIG, b, p))

def lin_blend(L: float, R: float, p: float, pL: float, pR: float) -> float:
    wL = (pR - p) / (pR - pL)
    wR = (p - pL) / (pR - pL)
    return wL * L + wR * R

def lin_blend_sigma(L: float, R: float, p: float, pL: float, pR: float) -> float:
    # same weights for σ(p)
    return lin_blend(L, R, p, pL, pR)

def M3_pairwise_Q(p: float) -> float:
    # Piecewise:
    # p<=1/12: BD
    # 1/12..1/6: BD->AE
    # 1/6..1/4: AE->AD
    # 1/4..7/12: AD
    # 7/12..2/3: BD->AB  (as per earlier construct)
    # 2/3..3/4: AB->BE
    # >=3/4: BE
    if p <= 1/12:
        return P_pair(PAIR_BD, p)
    if p < 1/6:
        return lin_blend(P_pair(PAIR_BD, p), P_pair(PAIR_AE, p), p, 1/12, 1/6)
    if p < 1/4:
        return lin_blend(P_pair(PAIR_AE, p), P_pair(PAIR_AD, p), p, 1/6, 1/4)
    if p < 7/12:
        return P_pair(PAIR_AD, p)
    if p < 2/3:
        return lin_blend(P_pair(PAIR_BD, p), P_pair(PAIR_AB, p), p, 7/12, 2/3)
    if p < 3/4:
        return lin_blend(P_pair(PAIR_AB, p), P_pair(PAIR_BE, p), p, 2/3, 3/4)
    return P_pair(PAIR_BE, p)

def M3_pairwise_sigma(p: float) -> float:
    if p <= 1/12:
        return P_pair_sigma(PAIR_BD, p)
    if p < 1/6:
        return lin_blend_sigma(P_pair_sigma(PAIR_BD, p), P_pair_sigma(PAIR_AE, p), p, 1/12, 1/6)
    if p < 1/4:
        return lin_blend_sigma(P_pair_sigma(PAIR_AE, p), P_pair_sigma(PAIR_AD, p), p, 1/6, 1/4)
    if p < 7/12:
        return P_pair_sigma(PAIR_AD, p)
    if p < 2/3:
        return lin_blend_sigma(P_pair_sigma(PAIR_BD, p), P_pair_sigma(PAIR_AB, p), p, 7/12, 2/3)
    if p < 3/4:
        return lin_blend_sigma(P_pair_sigma(PAIR_AB, p), P_pair_sigma(PAIR_BE, p), p, 2/3, 3/4)
    return P_pair_sigma(PAIR_BE, p)

def M1_DJ_Q(p: float) -> float:
    # Your M1 frame (with normalization fixed): use A..E..B structure
    # D ≤ 1/4 < 12(D(1/3-p)+A(p-1/4)) < 1/3 ≤ 3(A(2/3-p)+E(p-1/3)) ≤ 2/3 < 12(E(3/4-p)+B(p-2/3)) < 3/4 ≤ B
    if p <= 1/4:
        return Q_system(AVG_SIG, "D", p)
    if p < 1/3:
        wD = 12.0 * (1/3 - p)
        wA = 12.0 * (p - 1/4)
        return wD * Q_system(AVG_SIG, "D", p) + wA * Q_system(AVG_SIG, "A", p)
    if p <= 2/3:
        wA = 3.0 * (2/3 - p)
        wE = 3.0 * (p - 1/3)
        return wA * Q_system(AVG_SIG, "A", p) + wE * Q_system(AVG_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * Q_system(AVG_SIG, "E", p) + wB * Q_system(AVG_SIG, "B", p)
    return Q_system(AVG_SIG, "B", p)

def M1_DJ_sigma(p: float) -> float:
    if p <= 1/4:
        return sigma_side(AVG_SIG, "D", p)
    if p < 1/3:
        wD = 12.0 * (1/3 - p)
        wA = 12.0 * (p - 1/4)
        return wD * sigma_side(AVG_SIG, "D", p) + wA * sigma_side(AVG_SIG, "A", p)
    if p <= 2/3:
        wA = 3.0 * (2/3 - p)
        wE = 3.0 * (p - 1/3)
        return wA * sigma_side(AVG_SIG, "A", p) + wE * sigma_side(AVG_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * sigma_side(AVG_SIG, "E", p) + wB * sigma_side(AVG_SIG, "B", p)
    return sigma_side(AVG_SIG, "B", p)

def M2_DJ_Q(p: float) -> float:
    # Mean-of-stacks: average Dunbar composite and Janson composite quantiles
    return 0.5 * (Dunbar_composite_Q(p) + Janson_composite_Q(p))

def M2_DJ_sigma(p: float) -> float:
    return 0.5 * (Dunbar_composite_sigma(p) + Janson_composite_sigma(p))

def M4_mean_of_means_Q(p: float) -> float:
    return 0.5 * (M1_DJ_Q(p) + M2_DJ_Q(p))

def M4_mean_of_means_sigma(p: float) -> float:
    return 0.5 * (M1_DJ_sigma(p) + M2_DJ_sigma(p))

def M5_triple_mean_Q(p: float) -> float:
    return (M1_DJ_Q(p) + M2_DJ_Q(p) + M3_pairwise_Q(p)) / 3.0

def M5_triple_mean_sigma(p: float) -> float:
    return (M1_DJ_sigma(p) + M2_DJ_sigma(p) + M3_pairwise_sigma(p)) / 3.0

# M6: winner-stitched with REQUIRED anchors to M3 at {1/6, 5/24, 7/12, 2/3}
# plus knot winners at specific p’s (from our winner map),
# smoothed linearly between adjacent knots.
M6_KNOTS: List[Tuple[float, str]] = [
    (1/12, "M3"),
    (1/6,  "M3"),   # hard anchor
    (5/24, "M3"),   # hard anchor
    (1/4,  "M4"),
    (7/24, "M5"),
    (1/3,  "M2"),
    (5/12, "M3"),
    (1/2,  "M3"),
    (7/12, "M3"),   # hard anchor
    (2/3,  "M3"),   # hard anchor
    (17/24,"M2"),
    (3/4,  "M3"),
    (11/12,"M5"),
]

def _get_model_Q(name: str) -> Callable[[float], float]:
    return {
        "M1": M1_DJ_Q,
        "M2": M2_DJ_Q,
        "M3": M3_pairwise_Q,
        "M4": M4_mean_of_means_Q,
        "M5": M5_triple_mean_Q,
    }[name]

def _get_model_sigma(name: str) -> Callable[[float], float]:
    return {
        "M1": M1_DJ_sigma,
        "M2": M2_DJ_sigma,
        "M3": M3_pairwise_sigma,
        "M4": M4_mean_of_means_sigma,
        "M5": M5_triple_mean_sigma,
    }[name]

def M6_Q(p: float) -> float:
    # exact at knots, linear blend between knots
    for pk, mk in M6_KNOTS:
        if abs(p - pk) < 1e-12:
            return _get_model_Q(mk)(p)
    # find bracket
    for i in range(len(M6_KNOTS) - 1):
        pL, mL = M6_KNOTS[i]
        pR, mR = M6_KNOTS[i + 1]
        if pL < p < pR:
            return lin_blend(_get_model_Q(mL)(p), _get_model_Q(mR)(p), p, pL, pR)
    # outside: clamp
    if p < M6_KNOTS[0][0]:
        return _get_model_Q(M6_KNOTS[0][1])(p)
    return _get_model_Q(M6_KNOTS[-1][1])(p)

def M6_sigma(p: float) -> float:
    for pk, mk in M6_KNOTS:
        if abs(p - pk) < 1e-12:
            return _get_model_sigma(mk)(p)
    for i in range(len(M6_KNOTS) - 1):
        pL, mL = M6_KNOTS[i]
        pR, mR = M6_KNOTS[i + 1]
        if pL < p < pR:
            return lin_blend_sigma(_get_model_sigma(mL)(p), _get_model_sigma(mR)(p), p, pL, pR)
    if p < M6_KNOTS[0][0]:
        return _get_model_sigma(M6_KNOTS[0][1])(p)
    return _get_model_sigma(M6_KNOTS[-1][1])(p)


# -----------------------------
# Inversion: N -> p for monotone Q(p)
# -----------------------------

def invert_Q_to_p(Qfunc: Callable[[float], float], target_N: float, lo=1e-12, hi=1-1e-12, iters=90) -> float:
    for _ in range(iters):
        mid = (lo + hi) / 2.0
        val = Qfunc(mid)
        if val < target_N:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2.0


# -----------------------------
# Workbook helpers
# -----------------------------

HEADER_FILL = PatternFill("solid", fgColor="EEECE1")

def setup_sheet(ws, title: str):
    ws.freeze_panes = "A3"
    ws["A1"] = title
    ws.merge_cells("A1:O1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "N",
        "p_F", "z_F", "sigma_F",
        "p_model", "z_model", "sigma_model",
        "Δp (model-F)", "Δz (model-F)", "Δsigma (model-F)",
        "N_F(p_model)", "N_model(p_model)", "ΔN(p_model)",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    widths = [8, 16, 12, 12, 16, 12, 14, 14, 14, 16, 14, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_row(ws, row):
    ws.append(row)

def fnum(x: float, ndp: int = 12) -> float:
    # keep numeric; Excel formatting can round
    return float(f"{x:.{ndp}g}")


# -----------------------------
# Main build
# -----------------------------

def build_sheet_for_regime(
    ws,
    title: str,
    fitF: SkewNormalFit,
    Q_model: Callable[[float], float],
    sigma_model: Callable[[float], float],
):
    setup_sheet(ws, title)

    for N in range(N_MIN, N_MAX + 1):
        # System F (skew-normal)
        pF = fitF.cdf(N)
        zF = z_std(pF)
        sigF = fitF.omega

        # model p by inversion (solve Q_model(p)=N)
        pM = invert_Q_to_p(Q_model, float(N))
        zM = z_std(pM)
        sigM = sigma_model(pM)

        # differences
        dp = pM - pF
        dz = zM - zF
        ds = sigM - sigF

        # Compare N at same p (p_model): how far F differs from model at that p
        N_F_at_pM = fitF.inv_cdf(pM)
        N_M_at_pM = Q_model(pM)
        dN = N_M_at_pM - N_F_at_pM

        write_row(ws, [
            N,
            fnum(pF, 15), fnum(zF, 12), fnum(sigF, 12),
            fnum(pM, 15), fnum(zM, 12), fnum(sigM, 12),
            fnum(dp, 15), fnum(dz, 12), fnum(ds, 12),
            fnum(N_F_at_pM, 12), fnum(N_M_at_pM, 12), fnum(dN, 12),
        ])

def build_definitions_sheet(ws):
    ws["A1"] = "Definitions & Equations"
    ws["A1"].font = Font(bold=True, size=14)

    lines = []
    lines.append("Common:")
    lines.append(f"  μ = {MU}")
    lines.append("  z(p) = Φ^{-1}(p) where Φ is the standard normal CDF.")
    lines.append("")
    lines.append("System F (skew-normal):")
    lines.append("  Standard skew-normal: f(z)=2 φ(z) Φ(α z);  F(z)=Φ(z) − 2*T(z,α) where T is Owen’s T.")
    lines.append("  X = ξ + ω Z, Z ~ SN(0,1,α).")
    lines.append("  Mean constraint: E[X]=μ => ξ = μ − ω*δ*sqrt(2/pi), δ = α/sqrt(1+α^2).")
    lines.append("  Fit solves: Q(0.025)=L and Q(0.975)=U with mean fixed at μ.")
    lines.append("  Reported sigma_F = ω (constant per fitted regime).")
    lines.append("")
    lines.append("Composite fuzzy logic (Dunbar):")
    lines.append("  p ≤ 1/4: C")
    lines.append("  1/4 < p < 1/3: 12*( C*(1/3−p) + D*(p−1/4) )")
    lines.append("  1/3 ≤ p < 1/2: 6*( D*(1/2−p) + E*(p−1/3) )")
    lines.append("  p ≥ 1/2: E")
    lines.append("")
    lines.append("Composite fuzzy logic (Janson):")
    lines.append("  p ≤ 1/6: D")
    lines.append("  1/6 < p < 1/4: 12*( D*(1/4−p) + A*(p−1/6) )")
    lines.append("  1/4 ≤ p < 1/3: A")
    lines.append("  1/3 ≤ p < 1/2: 6*( A*(1/2−p) + E*(p−1/3) )")
    lines.append("  1/2 ≤ p ≤ 2/3: E")
    lines.append("  2/3 < p < 3/4: 12*( E*(3/4−p) + B*(p−2/3) )")
    lines.append("  p ≥ 3/4: B")
    lines.append("")
    lines.append("Dunbar-Janson mean regime models:")
    lines.append("  AVG σ parameters are componentwise means of Dunbar and Janson σ’s.")
    lines.append("  M1_DJ uses D/A/E/B stack (with normalization at [1/3,2/3]: factor 3).")
    lines.append("  M2_DJ = 0.5*(Dunbar composite + Janson composite).")
    lines.append("  M3 = pairwise-constructed (BD, AE, AD, AB, BE) with linear smoothing between knots.")
    lines.append("  M4 = 0.5*(M1_DJ + M2_DJ).")
    lines.append("  M5 = (M1_DJ + M2_DJ + M3)/3.")
    lines.append("  M6 = winner-stitched knots (hard anchors to M3 at p={1/6,5/24,7/12,2/3}) with linear smoothing between knots.")
    lines.append("")
    lines.append("σ(p) reporting:")
    lines.append("  For A,C,D: σ is constant.")
    lines.append("  For B,E: σ(p)=σ_L if p<0.5 else σ_U.")
    lines.append("  For fuzzy blends: σ(p) blends with the same weights as the quantiles.")
    lines.append("")
    lines.append("Difference columns:")
    lines.append("  Δp = p_model − p_F; Δz = z_model − z_F; Δsigma = sigma_model − sigma_F.")
    lines.append("  ΔN(p_model) compares the model quantile at p_model to F’s quantile at p_model.")

    ws["A3"] = "\n".join(lines)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 140
    ws.row_dimensions[3].height = 900


def main():
    print("Fitting skew-normal System F parameters...")
    fit_dunbar = fit_skewnormal_from_bounds(DUNBAR_L, DUNBAR_U)
    fit_janson = fit_skewnormal_from_bounds(JANSON_L, JANSON_U)
    fit_dj = fit_skewnormal_from_bounds(DJ_L, DJ_U)

    print("Dunbar F:", fit_dunbar)
    print("Janson F:", fit_janson)
    print("DJ F:", fit_dj)

    wb = Workbook()
    # remove default
    wb.remove(wb.active)

    # Dunbar sheet
    ws_d = wb.create_sheet("Dunbar")
    build_sheet_for_regime(
        ws_d,
        title="Dunbar: System F (skew-normal fit) vs Composite Fuzzy; differences",
        fitF=fit_dunbar,
        Q_model=Dunbar_composite_Q,
        sigma_model=Dunbar_composite_sigma,
    )

    # Janson sheet
    ws_j = wb.create_sheet("Janson")
    build_sheet_for_regime(
        ws_j,
        title="Janson: System F (skew-normal fit) vs Composite Fuzzy; differences",
        fitF=fit_janson,
        Q_model=Janson_composite_Q,
        sigma_model=Janson_composite_sigma,
    )

    # Dunbar-Janson sheet: F vs M6
    ws_dj = wb.create_sheet("Dunbar-Janson")
    build_sheet_for_regime(
        ws_dj,
        title="Dunbar–Janson: System F (skew-normal fit to averaged bounds) vs M6; differences",
        fitF=fit_dj,
        Q_model=M6_Q,
        sigma_model=M6_sigma,
    )

    # Definitions
    ws_def = wb.create_sheet("Definitions")
    build_definitions_sheet(ws_def)

    out = "Composite_Fuzzy_Dunbar_Janson_DJ_M6_N9-564.xlsx"
    wb.save(out)
    print("Saved:", out)


if __name__ == "__main__":
    main()
