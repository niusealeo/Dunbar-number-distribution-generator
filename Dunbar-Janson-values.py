#!/usr/bin/env python3
"""
Build an XLSX workbook for N = 9..564 with columns p, z, sigma, WITHOUT any skew-normal (System F).

Sheets:
- Dunbar: Composite fuzzy logic (your smoothed-around-0.5 equations)
- Janson: Composite fuzzy logic (your smoothed-around-0.5 equations)
- Dunbar-Janson: M6 (winner-stitched + fuzzy smoothing, with hard anchors to M3 at p={1/6,5/24,7/12,2/3})
- Definitions: equations, parameters, and interpretation

Conventions:
- mu is fixed at 147.8 for all models.
- z is standard normal z = Φ^{-1}(p).
- sigma is the model’s effective σ(p) from underlying A–E systems (B/E use σL for p<0.5 and σU for p≥0.5).
- p is obtained by inverting N = Q_model(p) using monotone bisection (no heavy numerics).

Requires:
    pip install openpyxl
"""

from __future__ import annotations

import math
from statistics import NormalDist
from typing import Callable, Dict, Tuple, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Global constants
# -----------------------------

MU = 147.8
ND = NormalDist()

N_MIN = 9
N_MAX = 564


# -----------------------------
# A–E σ parameters (pre-averaging)
# -----------------------------

DUNBAR_SIG = {
    "A": {"sigma": 33.393},
    "C": {"sigma": 28.629},
    "D": {"sigma": 31.011},
    "B": {"sigma_L": 24.286, "sigma_U": 42.501},
    "E": {"sigma_L": 26.458, "sigma_U": 35.565},
}

JANSON_SIG = {
    "A": {"sigma": 108.03004630194182},
    "C": {"sigma": 76.0177544515727},
    "D": {"sigma": 92.02390037675727},
    "B": {"sigma_L": 63.812397057566606, "sigma_U": 152.24769554631703},
    "E": {"sigma_L": 69.91507575456966, "sigma_U": 114.13272499894487},
}

# Dunbar–Janson averaged σ (componentwise mean)
AVG_SIG: Dict[str, Dict[str, float]] = {}
for k in ["A", "C", "D"]:
    AVG_SIG[k] = {"sigma": (DUNBAR_SIG[k]["sigma"] + JANSON_SIG[k]["sigma"]) / 2.0}
for k in ["B", "E"]:
    AVG_SIG[k] = {
        "sigma_L": (DUNBAR_SIG[k]["sigma_L"] + JANSON_SIG[k]["sigma_L"]) / 2.0,
        "sigma_U": (DUNBAR_SIG[k]["sigma_U"] + JANSON_SIG[k]["sigma_U"]) / 2.0,
    }


# -----------------------------
# Helpers: Normal quantile model Q(p)=MU+sigma(p)*z(p)
# -----------------------------

def z_std(p: float) -> float:
    return ND.inv_cdf(p)

def sigma_side(sig_params: Dict, sys: str, p: float) -> float:
    if sys in ("B", "E"):
        return sig_params[sys]["sigma_L"] if p < 0.5 else sig_params[sys]["sigma_U"]
    return sig_params[sys]["sigma"]

def Q_system(sig_params: Dict, sys: str, p: float) -> float:
    return MU + sigma_side(sig_params, sys, p) * z_std(p)


# -----------------------------
# Composite fuzzy logic (your updated smoothing around 0.5)
# -----------------------------

def Dunbar_composite_Q(p: float) -> float:
    # Dunbar frame:
    # C ≤ 1/4 < 12(C(1/3-p)+D(p-1/4)) < 1/3 ≤ 6(D(1/2-p)+E(p-1/3)) < 1/2 < E
    if p <= 1/4:
        return Q_system(DUNBAR_SIG, "C", p)
    if p < 1/3:
        wC = 12.0 * (1/3 - p)
        wD = 12.0 * (p - 1/4)
        return wC * Q_system(DUNBAR_SIG, "C", p) + wD * Q_system(DUNBAR_SIG, "D", p)
    if p < 1/2:
        wD = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wD * Q_system(DUNBAR_SIG, "D", p) + wE * Q_system(DUNBAR_SIG, "E", p)
    return Q_system(DUNBAR_SIG, "E", p)

def Dunbar_composite_sigma(p: float) -> float:
    if p <= 1/4:
        return sigma_side(DUNBAR_SIG, "C", p)
    if p < 1/3:
        wC = 12.0 * (1/3 - p)
        wD = 12.0 * (p - 1/4)
        return wC * sigma_side(DUNBAR_SIG, "C", p) + wD * sigma_side(DUNBAR_SIG, "D", p)
    if p < 1/2:
        wD = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wD * sigma_side(DUNBAR_SIG, "D", p) + wE * sigma_side(DUNBAR_SIG, "E", p)
    return sigma_side(DUNBAR_SIG, "E", p)

def Janson_composite_Q(p: float) -> float:
    # Janson frame:
    # D ≤ 1/6 < 12(D(1/4-p)+A(p-1/6)) < 1/4 ≤ 6(A(1/2-p)+E(p-1/3)) < 1/2 < E ≤ 2/3 < 12(E(3/4-p)+B(p-2/3)) < 3/4 ≤ B
    if p <= 1/6:
        return Q_system(JANSON_SIG, "D", p)
    if p < 1/4:
        wD = 12.0 * (1/4 - p)
        wA = 12.0 * (p - 1/6)
        return wD * Q_system(JANSON_SIG, "D", p) + wA * Q_system(JANSON_SIG, "A", p)
    if p < 1/3:
        return Q_system(JANSON_SIG, "A", p)
    if p < 1/2:
        wA = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wA * Q_system(JANSON_SIG, "A", p) + wE * Q_system(JANSON_SIG, "E", p)
    if p <= 2/3:
        return Q_system(JANSON_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * Q_system(JANSON_SIG, "E", p) + wB * Q_system(JANSON_SIG, "B", p)
    return Q_system(JANSON_SIG, "B", p)

def Janson_composite_sigma(p: float) -> float:
    if p <= 1/6:
        return sigma_side(JANSON_SIG, "D", p)
    if p < 1/4:
        wD = 12.0 * (1/4 - p)
        wA = 12.0 * (p - 1/6)
        return wD * sigma_side(JANSON_SIG, "D", p) + wA * sigma_side(JANSON_SIG, "A", p)
    if p < 1/3:
        return sigma_side(JANSON_SIG, "A", p)
    if p < 1/2:
        wA = 6.0 * (1/2 - p)
        wE = 6.0 * (p - 1/3)
        return wA * sigma_side(JANSON_SIG, "A", p) + wE * sigma_side(JANSON_SIG, "E", p)
    if p <= 2/3:
        return sigma_side(JANSON_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * sigma_side(JANSON_SIG, "E", p) + wB * sigma_side(JANSON_SIG, "B", p)
    return sigma_side(JANSON_SIG, "B", p)


# -----------------------------
# Dunbar–Janson regime models for M6 (with latest renames)
# -----------------------------

# Pairwise-constructed model (NEW M3):
PAIR_BD = ("B", "D")
PAIR_AE = ("A", "E")
PAIR_AD = ("A", "D")
PAIR_AB = ("A", "B")
PAIR_BE = ("B", "E")

def P_pair(pair: Tuple[str, str], p: float) -> float:
    a, b = pair
    return 0.5 * (Q_system(AVG_SIG, a, p) + Q_system(AVG_SIG, b, p))

def P_pair_sigma(pair: Tuple[str, str], p: float) -> float:
    a, b = pair
    return 0.5 * (sigma_side(AVG_SIG, a, p) + sigma_side(AVG_SIG, b, p))

def lin_blend(L: float, R: float, p: float, pL: float, pR: float) -> float:
    wL = (pR - p) / (pR - pL)
    wR = (p - pL) / (pR - pL)
    return wL * L + wR * R

def M3_pairwise_Q(p: float) -> float:
    if p <= 1/12:
        return P_pair(PAIR_BD, p)
    if p < 1/6:
        return lin_blend(P_pair(PAIR_BD, p), P_pair(PAIR_AE, p), p, 1/12, 1/6)
    if p < 1/4:
        return lin_blend(P_pair(PAIR_AE, p), P_pair(PAIR_AD, p), p, 1/6, 1/4)
    if p < 7/12:
        return P_pair(PAIR_AD, p)
    if p < 2/3:
        return lin_blend(P_pair(PAIR_BD, p), P_pair(PAIR_AB, p), p, 7/12, 2/3)
    if p < 3/4:
        return lin_blend(P_pair(PAIR_AB, p), P_pair(PAIR_BE, p), p, 2/3, 3/4)
    return P_pair(PAIR_BE, p)

def M3_pairwise_sigma(p: float) -> float:
    if p <= 1/12:
        return P_pair_sigma(PAIR_BD, p)
    if p < 1/6:
        return lin_blend(P_pair_sigma(PAIR_BD, p), P_pair_sigma(PAIR_AE, p), p, 1/12, 1/6)
    if p < 1/4:
        return lin_blend(P_pair_sigma(PAIR_AE, p), P_pair_sigma(PAIR_AD, p), p, 1/6, 1/4)
    if p < 7/12:
        return P_pair_sigma(PAIR_AD, p)
    if p < 2/3:
        return lin_blend(P_pair_sigma(PAIR_BD, p), P_pair_sigma(PAIR_AB, p), p, 7/12, 2/3)
    if p < 3/4:
        return lin_blend(P_pair_sigma(PAIR_AB, p), P_pair_sigma(PAIR_BE, p), p, 2/3, 3/4)
    return P_pair_sigma(PAIR_BE, p)

def M1_DJ_Q(p: float) -> float:
    # M1 frame (normalised):
    # D ≤ 1/4 < 12(D(1/3-p)+A(p-1/4)) < 1/3 ≤ 3(A(2/3-p)+E(p-1/3)) ≤ 2/3 < 12(E(3/4-p)+B(p-2/3)) < 3/4 ≤ B
    if p <= 1/4:
        return Q_system(AVG_SIG, "D", p)
    if p < 1/3:
        wD = 12.0 * (1/3 - p)
        wA = 12.0 * (p - 1/4)
        return wD * Q_system(AVG_SIG, "D", p) + wA * Q_system(AVG_SIG, "A", p)
    if p <= 2/3:
        wA = 3.0 * (2/3 - p)
        wE = 3.0 * (p - 1/3)
        return wA * Q_system(AVG_SIG, "A", p) + wE * Q_system(AVG_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * Q_system(AVG_SIG, "E", p) + wB * Q_system(AVG_SIG, "B", p)
    return Q_system(AVG_SIG, "B", p)

def M1_DJ_sigma(p: float) -> float:
    if p <= 1/4:
        return sigma_side(AVG_SIG, "D", p)
    if p < 1/3:
        wD = 12.0 * (1/3 - p)
        wA = 12.0 * (p - 1/4)
        return wD * sigma_side(AVG_SIG, "D", p) + wA * sigma_side(AVG_SIG, "A", p)
    if p <= 2/3:
        wA = 3.0 * (2/3 - p)
        wE = 3.0 * (p - 1/3)
        return wA * sigma_side(AVG_SIG, "A", p) + wE * sigma_side(AVG_SIG, "E", p)
    if p < 3/4:
        wE = 12.0 * (3/4 - p)
        wB = 12.0 * (p - 2/3)
        return wE * sigma_side(AVG_SIG, "E", p) + wB * sigma_side(AVG_SIG, "B", p)
    return sigma_side(AVG_SIG, "B", p)

def M2_DJ_Q(p: float) -> float:
    return 0.5 * (Dunbar_composite_Q(p) + Janson_composite_Q(p))

def M2_DJ_sigma(p: float) -> float:
    return 0.5 * (Dunbar_composite_sigma(p) + Janson_composite_sigma(p))

def M4_Q(p: float) -> float:
    return 0.5 * (M1_DJ_Q(p) + M2_DJ_Q(p))

def M4_sigma(p: float) -> float:
    return 0.5 * (M1_DJ_sigma(p) + M2_DJ_sigma(p))

def M5_Q(p: float) -> float:
    return (M1_DJ_Q(p) + M2_DJ_Q(p) + M3_pairwise_Q(p)) / 3.0

def M5_sigma(p: float) -> float:
    return (M1_DJ_sigma(p) + M2_DJ_sigma(p) + M3_pairwise_sigma(p)) / 3.0

# Corrected M6: explicit knots and anchors (as agreed)
M6_KNOTS: List[Tuple[float, str]] = [
    (1/12, "M3"),
    (1/6,  "M3"),   # hard anchor
    (5/24, "M3"),   # hard anchor
    (1/4,  "M4"),
    (7/24, "M5"),
    (1/3,  "M2"),
    (5/12, "M3"),
    (1/2,  "M3"),
    (7/12, "M3"),   # hard anchor
    (2/3,  "M3"),   # hard anchor
    (17/24,"M2"),
    (3/4,  "M3"),
    (11/12,"M5"),
]

def _get_Q(name: str) -> Callable[[float], float]:
    return {"M1": M1_DJ_Q, "M2": M2_DJ_Q, "M3": M3_pairwise_Q, "M4": M4_Q, "M5": M5_Q}[name]

def _get_sigma(name: str) -> Callable[[float], float]:
    return {"M1": M1_DJ_sigma, "M2": M2_DJ_sigma, "M3": M3_pairwise_sigma, "M4": M4_sigma, "M5": M5_sigma}[name]

def M6_Q(p: float) -> float:
    for pk, mk in M6_KNOTS:
        if abs(p - pk) < 1e-12:
            return _get_Q(mk)(p)
    for i in range(len(M6_KNOTS) - 1):
        pL, mL = M6_KNOTS[i]
        pR, mR = M6_KNOTS[i + 1]
        if pL < p < pR:
            return lin_blend(_get_Q(mL)(p), _get_Q(mR)(p), p, pL, pR)
    if p < M6_KNOTS[0][0]:
        return _get_Q(M6_KNOTS[0][1])(p)
    return _get_Q(M6_KNOTS[-1][1])(p)

def M6_sigma(p: float) -> float:
    for pk, mk in M6_KNOTS:
        if abs(p - pk) < 1e-12:
            return _get_sigma(mk)(p)
    for i in range(len(M6_KNOTS) - 1):
        pL, mL = M6_KNOTS[i]
        pR, mR = M6_KNOTS[i + 1]
        if pL < p < pR:
            return lin_blend(_get_sigma(mL)(p), _get_sigma(mR)(p), p, pL, pR)
    if p < M6_KNOTS[0][0]:
        return _get_sigma(M6_KNOTS[0][1])(p)
    return _get_sigma(M6_KNOTS[-1][1])(p)


# -----------------------------
# Inversion: N -> p for monotone Q(p)
# -----------------------------

def invert_Q_to_p(Qfunc: Callable[[float], float], target_N: float, lo=1e-12, hi=1-1e-12, iters=90) -> float:
    for _ in range(iters):
        mid = (lo + hi) / 2.0
        val = Qfunc(mid)
        if val < target_N:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2.0


# -----------------------------
# Workbook formatting
# -----------------------------

HEADER_FILL = PatternFill("solid", fgColor="EEECE1")

def fnum(x: float, ndp: int = 15) -> float:
    return float(f"{x:.{ndp}g}")

def setup_sheet(ws, title: str):
    ws.freeze_panes = "A3"
    ws["A1"] = title
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["N", "p", "z", "sigma", "Q(p)", "check (Q(p)-N)"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    widths = [8, 18, 14, 14, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_simple_sheet(
    ws,
    title: str,
    Q_model: Callable[[float], float],
    sigma_model: Callable[[float], float],
):
    setup_sheet(ws, title)

    for N in range(N_MIN, N_MAX + 1):
        p = invert_Q_to_p(Q_model, float(N))
        z = z_std(p)
        sig = sigma_model(p)
        Qp = Q_model(p)
        chk = Qp - N
        ws.append([
            N,
            fnum(p, 15),
            fnum(z, 12),
            fnum(sig, 12),
            fnum(Qp, 12),
            fnum(chk, 12),
        ])

def build_definitions_sheet(ws):
    ws["A1"] = "Definitions & Equations (No skew-normal / no System F)"
    ws["A1"].font = Font(bold=True, size=14)

    lines = []
    lines.append("Common:")
    lines.append(f"  μ = {MU}")
    lines.append("  z(p) = Φ^{-1}(p) where Φ is the standard normal CDF.")
    lines.append("  Base system quantile: Q_X(p) = μ + σ_X(p)*z(p).")
    lines.append("")
    lines.append("σ(p) conventions:")
    lines.append("  For A,C,D: σ is constant.")
    lines.append("  For B,E: σ(p)=σ_L if p<0.5 else σ_U.")
    lines.append("  For fuzzy blends: σ(p) blends using the same weights as the quantiles.")
    lines.append("")
    lines.append("Composite fuzzy logic (Dunbar):")
    lines.append("  p ≤ 1/4: C")
    lines.append("  1/4 < p < 1/3: 12*( C*(1/3−p) + D*(p−1/4) )")
    lines.append("  1/3 ≤ p < 1/2: 6*( D*(1/2−p) + E*(p−1/3) )")
    lines.append("  p ≥ 1/2: E")
    lines.append("")
    lines.append("Composite fuzzy logic (Janson):")
    lines.append("  p ≤ 1/6: D")
    lines.append("  1/6 < p < 1/4: 12*( D*(1/4−p) + A*(p−1/6) )")
    lines.append("  1/4 ≤ p < 1/3: A")
    lines.append("  1/3 ≤ p < 1/2: 6*( A*(1/2−p) + E*(p−1/3) )")
    lines.append("  1/2 ≤ p ≤ 2/3: E")
    lines.append("  2/3 < p < 3/4: 12*( E*(3/4−p) + B*(p−2/3) )")
    lines.append("  p ≥ 3/4: B")
    lines.append("")
    lines.append("Dunbar–Janson averaged σ parameters:")
    lines.append("  AVG σ parameters are componentwise means of Dunbar and Janson σ’s.")
    lines.append("")
    lines.append("Pairwise-constructed model (M3):")
    lines.append("  Uses pairs (B+D), (A+E), (A+D), (A+B), (B+E) with linear smoothing between knots.")
    lines.append("")
    lines.append("M6 (winner-stitched + fuzzy smoothing):")
    lines.append("  Knots (p -> model):")
    for p, m in M6_KNOTS:
        lines.append(f"    p={p} -> {m}")
    lines.append("  M6(p) equals the knot model at knot p, and is linearly interpolated between adjacent knots.")
    lines.append("  Hard anchors to M3 are included at p={1/6, 5/24, 7/12, 2/3}.")
    lines.append("")
    lines.append("σ parameters used (pre-averaging):")
    lines.append("  Dunbar: σA=33.393; σC=28.629; σD=31.011; σBL=24.286; σBU=42.501; σEL=26.458; σEU=35.565")
    lines.append("  Janson: σA=108.03004630194182; σC=76.0177544515727; σD=92.02390037675727;")
    lines.append("         σBL=63.812397057566606; σBU=152.24769554631703; σEL=69.91507575456966; σEU=114.13272499894487")

    ws["A3"] = "\n".join(lines)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 140
    ws.row_dimensions[3].height = 950


def main():
    wb = Workbook()
    wb.remove(wb.active)

    ws_d = wb.create_sheet("Dunbar")
    build_simple_sheet(
        ws_d,
        title="Dunbar: Composite fuzzy logic (no System F) — N=9..564",
        Q_model=Dunbar_composite_Q,
        sigma_model=Dunbar_composite_sigma,
    )

    ws_j = wb.create_sheet("Janson")
    build_simple_sheet(
        ws_j,
        title="Janson: Composite fuzzy logic (no System F) — N=9..564",
        Q_model=Janson_composite_Q,
        sigma_model=Janson_composite_sigma,
    )

    ws_dj = wb.create_sheet("Dunbar-Janson")
    build_simple_sheet(
        ws_dj,
        title="Dunbar–Janson: M6 (no System F) — N=9..564",
        Q_model=M6_Q,
        sigma_model=M6_sigma,
    )

    ws_def = wb.create_sheet("Definitions")
    build_definitions_sheet(ws_def)

    out = "Composite_Fuzzy_Dunbar_Janson_M6_N9-564_NO_F.xlsx"
    wb.save(out)
    print("Saved:", out)


if __name__ == "__main__":
    main()
